import json
import requests
import urllib.request, urllib.parse, urllib.error,requests
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
from bs4 import BeautifulSoup
import matplotlib.pyplot as plt
weathers=[]
location_api=[]
avg=[]
avg2=[]
sign=0
#30010新竹市東區大學路1001號
###########################################################
city=["臺北市","新北市","基隆市","桃園市","新竹市","新竹縣","苗栗縣","臺中市","彰化縣","南投縣","雲林縣","嘉義市","嘉義縣","臺南市","高雄市","宜蘭縣","花蓮縣","台東縣","澎湖縣","金門縣","連江縣"]
url1="https://maps.googleapis.com/maps/api/geocode/json?"
url2="https://maps.googleapis.com/maps/api/directions/json?"
url3="https://opendata.cwb.gov.tw/"
api_key="AIzaSyAzuC0VZbcsvlE9QNbd8x6LRCqphXEncxs"
language1="zh-TW"
api_au='CWB-50B0B6D6-E515-41C0-BED7-19D2984A6C12'
cnames = {
'burlywood':            '#DEB887',
'cadetblue':            '#5F9EA0',
'chartreuse':           '#7FFF00',
'chocolate':            '#D2691E',
'coral':                '#FF7F50',
'cornflowerblue':       '#6495ED',
'cornsilk':             '#FFF8DC',
'crimson':              '#DC143C',
'cyan':                 '#00FFFF',
'darkblue':             '#00008B',
'darkcyan':             '#008B8B',
'darkgoldenrod':        '#B8860B',
'darkgray':             '#A9A9A9',
'darkgreen':            '#006400',
'darkkhaki':            '#BDB76B',
'darkmagenta':          '#8B008B',
'darkolivegreen':       '#556B2F',
'darkorange':           '#FF8C00',
'darkorchid':           '#9932CC',
'darkred':              '#8B0000',
'darksalmon':           '#E9967A',
'darkseagreen':         '#8FBC8F',
'darkslateblue':        '#483D8B',
'darkslategray':        '#2F4F4F',
'darkturquoise':        '#00CED1',
'darkviolet':           '#9400D3',
'deeppink':             '#FF1493',
'deepskyblue':          '#00BFFF',
'dimgray':              '#696969',
'dodgerblue':           '#1E90FF',
'firebrick':            '#B22222',
'floralwhite':          '#FFFAF0',
'forestgreen':          '#228B22',
'fuchsia':              '#FF00FF',
'gainsboro':            '#DCDCDC',
'ghostwhite':           '#F8F8FF',
'gold':                 '#FFD700',
'goldenrod':            '#DAA520',
'gray':                 '#808080',
'green':                '#008000',
'greenyellow':          '#ADFF2F',
'honeydew':             '#F0FFF0',
'hotpink':              '#FF69B4',
'indianred':            '#CD5C5C',
'indigo':               '#4B0082',
'ivory':                '#FFFFF0',
'khaki':                '#F0E68C',
'lavender':             '#E6E6FA',
'lavenderblush':        '#FFF0F5',
'lawngreen':            '#7CFC00',
'lemonchiffon':         '#FFFACD',
'lightblue':            '#ADD8E6',
'lightcoral':           '#F08080',
'lightcyan':            '#E0FFFF',
'lightgoldenrodyellow': '#FAFAD2',
'lightgreen':           '#90EE90',
'lightgray':            '#D3D3D3',
'lightpink':            '#FFB6C1',
'lightsalmon':          '#FFA07A',
'lightseagreen':        '#20B2AA',
'lightskyblue':         '#87CEFA',
'lightslategray':       '#778899',
'lightsteelblue':       '#B0C4DE',
'lightyellow':          '#FFFFE0',
'lime':                 '#00FF00',
'limegreen':            '#32CD32',
'linen':                '#FAF0E6',
'magenta':              '#FF00FF',
'maroon':               '#800000',
'mediumaquamarine':     '#66CDAA',
'mediumblue':           '#0000CD',
'mediumorchid':         '#BA55D3',
'mediumpurple':         '#9370DB',
'mediumseagreen':       '#3CB371',
'mediumslateblue':      '#7B68EE',
'mediumspringgreen':    '#00FA9A',
'mediumturquoise':      '#48D1CC',
'mediumvioletred':      '#C71585',
'midnightblue':         '#191970',
'mintcream':            '#F5FFFA',
'mistyrose':            '#FFE4E1',
'moccasin':             '#FFE4B5',
'navajowhite':          '#FFDEAD',
'navy':                 '#000080',
'oldlace':              '#FDF5E6',
'olive':                '#808000',
'olivedrab':            '#6B8E23',
'orange':               '#FFA500',
'orangered':            '#FF4500',
'orchid':               '#DA70D6',
'palegoldenrod':        '#EEE8AA',
'palegreen':            '#98FB98',
'paleturquoise':        '#AFEEEE',
'palevioletred':        '#DB7093',
'papayawhip':           '#FFEFD5',
'peachpuff':            '#FFDAB9',
'peru':                 '#CD853F',
'pink':                 '#FFC0CB',
'plum':                 '#DDA0DD',
'powderblue':           '#B0E0E6',
'purple':               '#800080',
'red':                  '#FF0000',
'rosybrown':            '#BC8F8F',
'royalblue':            '#4169E1',
'saddlebrown':          '#8B4513',
'salmon':               '#FA8072',
'sandybrown':           '#FAA460',
'seagreen':             '#2E8B57',
'seashell':             '#FFF5EE',
'sienna':               '#A0522D',
'silver':               '#C0C0C0',
'skyblue':              '#87CEEB',
'slateblue':            '#6A5ACD',
'slategray':            '#708090',
'snow':                 '#FFFAFA',
'springgreen':          '#00FF7F',
'steelblue':            '#4682B4',
'tan':                  '#D2B48C',
'teal':                 '#008080',
'thistle':              '#D8BFD8',
'tomato':               '#FF6347',
'turquoise':            '#40E0D0',
'violet':               '#EE82EE',
'wheat':                '#F5DEB3',
'white':                '#FFFFFF',
'whitesmoke':           '#F5F5F5',
'yellow':               '#FFFF00',
'yellowgreen':          '#9ACD32'}
color=[i for i in cnames]
###########################################################
def give_suggest():
    file=requests.get("https://gis.taiwan.net.tw/XMLReleaseALL_public/scenic_spot_C_f.json")
    jfile=file.json()
    want_info=jfile["XML_Head"]["Infos"]["Info"]
    location_ahs=[]
    introduction=[]
    for y in city:
        print(y)
    layer=1
    if (layer==1):
        while True:
            s=input("choose one city")
            if s in city:
                layer=2
                break
            elif(s=='b'):
                continue
            elif(s=='q'):
                break
            else:
                s=input("enter again")
    if (layer==2):
        for x in want_info:
            if s==x['Region']:
                location_ahs.append(x['Name'])
                introduction.append(x["Description"])
            else:
                continue
        layer=3
    if (layer==3):
        for y,z in enumerate(location_ahs):
            print(y,z)
        while True:
            try:
                num=int(input("choose one number and I will introduce it"))
                break
            except ValueError:
                continue
        print(location_ahs[num],":")
        print(introduction[num])
def get_geocode(location):
    file=requests.get(url1+"address="+location+"&language="+language1+"&key="+api_key)
    jfile=file.json()
    return jfile["results"][0]["place_id"]
def judge_id(id):
    sign=0
    x1=pd.ExcelFile("setting.xlsx")
    df=x1.parse("sheet1")
    row,column=df.shape
    fieldnames=df.keys()
    for irow in range(0,row):
        #print(df['ID'][irow])
        if(id==str(df["ID"][irow])):
            #print("ok")
            sign=1
            return irow+2
    if(sign==0):
        return -123
def create_id(id):
    row=ws1.max_row+1
    ws1.cell(row=ws1.max_row+1,column=1,value=id)
    #print(row)
    return row
    wb.save(filename='setting.xlsx')
def store_setting(setting,irow):
    if(setting=='y'):
        ws1.cell(row=irow,column=2,value=1)
        ws1.cell(row=irow,column=3,value=2)
    else:
        ws1.cell(row=irow,column=3,value=1)
        ws1.cell(row=irow,column=2,value=2)
    wb.save(filename='setting.xlsx')
def get_method_to_destionation_driving(ori_place_id,des_place_id):
    want_destr=[]
    file=requests.get(url2+"&origin=place_id:"+ori_place_id+"&language="+language1+"&destination=place_id:"+des_place_id+"&depature_time=now&mode=driving"+"&key="+api_key)
    jfile=file.json()
    for x in jfile['routes'][0]["legs"]:
        for y in x['steps']:
            #print(y['html_instructions'])
            a=y['html_instructions'].replace("<b>",'')
            b=a.replace("</b>",'')
            c=b.replace("/<wbr/>",'')
            d=c.replace("<div style=\"font-size:0.9em\">",'')
            e=d.replace("</div>","")
            f=e.replace("&nbsp;","")
            want_destr.append(f)
    distance=jfile["routes"][0]["legs"][0]['distance']['text']
    toll=distance.split("公里")[0]
    real_tolls=float(toll)*4
    time=jfile['routes'][0]['legs'][0]['duration']['text']
    real_time=time.split("分鐘")[0]
    if ('小時'in real_time):
        hour=float(real_time.split('小時')[0])
        minute=float(real_time.split('小時')[0])
        times=hour*60+minute
    else:
        times=float(real_time)
    # print(real_tolls)
    # print(times)
    return real_tolls,times,want_destr
def get_method_to_destionation_transit(ori_place_id,des_place_id):
    file=requests.get(url2+"&origin=place_id:"+ori_place_id+"&language="+language1+"&destination=place_id:"+des_place_id+"&depature_time=now&mode=transit"+"&key="+api_key)
    jfile=file.json()
    want_transit_destr=[]
    for x in jfile['routes'][0]["legs"]:
        for y in x['steps']:
            #print(y['html_instructions'])
            a=y['html_instructions'].replace("<b>",'')
            b=a.replace("</b>",'')
            c=b.replace("/<wbr/>",'')
            d=c.replace("<div style=\"font-size:0.9em\">",'')
            e=d.replace("</div>","")
            f=e.replace("&nbsp;","")
            want_transit_destr.append(f)
    #print(jfile)
    toll=jfile['routes'][0]['fare']['text']
    real_toll=toll.split("$")[1]
    time=jfile['routes'][0]['legs'][0]['duration']['text']
    real_time=time.split("分鐘")[0]
    if ('小時'in real_time):
        hour=float(real_time.split('小時')[0])
        minute=float(real_time.split('小時')[0])
        times=hour*60+minute
    else:
        times=float(real_time)
    # print(float(real_toll))
    # print(times)
    return float(real_toll),times,want_transit_destr
def get_weather_condition(want_location):
    picture_min=[]
    counts=0
    picture_y_min=[[y for y in range(len(want_location))] for x in range(3)]
    picture_x=[6,15,24]
    picture_x_min=[[x for y in range(len(want_location))]for x in picture_x]
    file=requests.get(url3+"/api/v1/rest/datastore/F-C0032-001?Authorization="+api_au+"&format=JSON")
    jfile=file.json()
    data=[[x]for x in range(len(want_location))]
   # print(data)
    #print(want_location)
    for z in range(len(want_location)):
        data[z][0]=want_location[z]
        for x in jfile['records']['location']:
            if(x['locationName']==want_location[z]):
                for y in x['weatherElement'][1]['time']:
                    data[z].append(y['parameter']['parameterName'])
                for a in x['weatherElement'][2]['time']:
                    picture_min.append(a['parameter']['parameterName'])
                for b in x['weatherElement'][4]['time']:
                    data[z].append(b['parameter']['parameterName'])
    
    #print(data)
    df=pd.DataFrame(data,columns=["region_name",'pop(12-18)','pop(18-06)','pop(06-18)','maxT(06-18)','maxT(18-06)','maxT(06-18)']) 
    print(df)
    for x in range(len(want_location)):
        for y in range(3):
            picture_y_min[y][x]=float(picture_min[counts])
            counts=counts+1
    counts=0
    plt.gca().set_prop_cycle('color', color)
    for x in range(len(want_location)):
        print(want_location[x]+" is "+color[x],'\'s color line')
    print("this is the picture about the miniuim temperature")
    plt.plot(picture_x_min,picture_y_min,linestyle='solid')
    plt.show()
    want_location.clear()
def scratch_news(reply):
    #print('ok')
    file=urllib.request.urlopen("https://1968.freeway.gov.tw/tf_ranking")
    d=file.read().decode()
    soup=BeautifulSoup(d,'html.parser')
    a=soup.find_all("div",class_="page _event_mmenu_add _event_smartbanner_add _event_map_expand")
    b=a[0].find_all("li",limit=11)
    for x,y in enumerate(b):
        print(x+1,",",y.text)
###########################################################
#  main function
###########################################################
layer=0
reply_setting_piority=[0,0]
while True:
    if(layer==0):
        reply_create=input("if you have the \"setting.xlsx\" in your computer?(input yes or no)")
        if(reply_create=='b'):
            layer=0
            continue
        elif(reply_create=='q'):
            break
        elif(reply_create=='yes'):
            wb=load_workbook(filename="setting.xlsx",data_only=True)
            ws1=wb['sheet1']
            ws1['A1']='ID'
            ws1['B1']='money'
            ws1['C1']='time'
            layer=1
        elif(reply_create=='no'):
            wa=Workbook()
            ws2=wa.active
            ws2.title='sheet1'
            ws2['A1']='ID'
            ws2['B1']='money'
            ws2['C1']='time'
            layer=0
            wa.save(filename="setting.xlsx")
            print("we create successful")
            continue
        else:
            print("input the data again")
            continue
    if(layer==1):
        reply_id=input("input the id:(b is back and q is leave): ")
        if(reply_id=='b'):
            layer=0
            continue
        if(reply_id=='q'):
            break
        if(judge_id(reply_id)==-123):
            reply_create_id=input("create a new id:(b is back and q is leave): ")
            if(reply_create_id=='b'):
                layer=1
                continue
            elif(reply_create_id=='q'):
                break
            iden=create_id(reply_create_id)
            print("successfully create new id: "+reply_create_id)
            layer=2
        else:
            #print("ok")
            iden=judge_id(reply_id)
            #print(iden)
            layer=100
##############################################################################
    if(layer==2):
        for x in range(2):
            reply_setting_piority[x]=input("please input the piority(b is back and q is leave)(we only take x:time, y:money into consideration): ")
            if(reply_setting_piority[x]=='b'):
                layer=1
                break
            if(reply_setting_piority[x]=='q'):
                break
            if(reply_setting_piority[x]!='x' and reply_setting_piority[x]!='y'):
                while True:
                    reply_setting_piority[x]=input("Error message!input again: ")
                    if(reply_setting_piority[x]=='x' or reply_setting_piority[x]=='y'):
                        break
        if(reply_setting_piority[0]=='q' or reply_setting_piority[1]=='q'):
            break
        if(layer==2):
            store_setting(reply_setting_piority[0],iden)
            layer=100
##############################################################################
    if(layer==100):
        judge=input("do you need some suggest where to go?(input yes or no)")
        if (judge=='yes'):
            give_suggest()
            layer=3
        elif(judge=='no'):
            layer=3
        elif(judge=='q'):
            break
        elif(judge=='b'):
            layer=2
            continue
        else:
            judge=input("enter again: ")
    if(layer==3):
        reply_change=input("do you want to change your setting?")
        if(reply_change=='yes'):
            layer=2
            continue
        elif(reply_change=='b'):
            layer=1
            continue
        elif(reply_change=='q'):
            break
        replyo=input("input the origin?(b is back and q is leave) ")
        if(replyo=='b'):
            layer=1
            continue
        elif(replyo=='q'):
             break
        replyd=input("input the destination you want to go ?(b is back and q is leave)")
        if(replyd=='b'):
          layer=3
          continue
        elif(replyd=='q'):
          break
        else:
            ori_geo_code_placeid=get_geocode(replyo)
            des_geo_code_placeid=get_geocode(replyd)
            driving_toll,driving_time,ans_driving_destr=get_method_to_destionation_driving(ori_geo_code_placeid,des_geo_code_placeid)
            transit_toll,transit_time,ans_transit_destr=get_method_to_destionation_transit(ori_geo_code_placeid,des_geo_code_placeid)
            layer=4
##############################################################################
    if(layer==4):
        if((ws1.cell(row=iden,column=2).value)<(ws1.cell(row=iden,column=3).value)):
            if(driving_toll<transit_toll):
                print("you should driving which costs ",driving_toll,"dollars and takes ",driving_time," mins")
                print("this is destrution:")
                for x in ans_driving_destr:
                    print(x)
                print("and there is the new about highway:")
                scratch_news
            elif(driving_toll==transit_toll):
                if(driving_time<transit_time):
                    print("you should driving which costs ",driving_toll,"dollars and takes ",driving_time," mins")
                    print("this is destrution:")
                    for x in ans_driving_destr:
                        print(x)
                    print("and there is the new about highway:")
                    scratch_news
                else:
                    print("you should take transit which costs ",transit_toll,"dollars and takes ",transit_time," mins")
                    print("this is destrution:")
                    for x in ans_transit_destr:
                        print(x)
            else:
                print("you should take transit which costs ",transit_toll,"dollars and takes ",transit_time," mins")
                print("this is destrution:")
                for x in ans_transit_destr:
                    print(x)
        else:
            if(driving_time<transit_time):
                print("you should driving which costs ",driving_toll,"dollars and takes ",driving_time," mins")
                print("this is destrution:")
                for x in ans_driving_destr:
                    print(x)
                print("and there is the new about highway:")
                scratch_news(iden)
            elif(driving_time==transit_time):
                if(driving_toll<transit_toll):
                    print("you should take transit which costs ",transit_toll,"dollars and takes ",transit_time," mins")
                    print("this is destrution:")
                    for x in ans_transit_destr:
                        print(x)
                else:
                    print("you should driving which costs ",driving_toll,"dollars and takes ",driving_time," mins")
                    print("this is destrution:")
                    for x in ans_driving_destr:
                        print(x)                    
                    print("and there is the new about highway:")
                    scratch_news(iden)
            else:
                print("you should take transit which costs ",transit_toll,"dollars and takes ",transit_time," mins")
                print("this is destrution:")
                for x in ans_transit_destr:
                    print(x)
        while True:
            weather=input('where\'s weather do you want to know?(if you input done then we finish this question)')
            if(weather=='b'):
                layer=3
                break
            elif(weather=='q'):
                break
            elif(weather=='done'):
                break
            elif(weather in city):
                print("you input successfully")
                sign=1
            else:
                print("please input again")
            if (sign==1):
                sign=0
                weathers.append(weather)
        if(weather=='q'):
            break
        get_weather_condition(weathers)
        layer=1
##############################################################################
